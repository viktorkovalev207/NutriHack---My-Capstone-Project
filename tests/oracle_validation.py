"""
INDEPENDENT ORACLE VALIDATION (the real bulletproof proof).

The workbook ships with only 3 worked example products — far too few to trust.
So we build our own oracle WITHOUT re-implementing the logic ourselves:

  1. Take the VERBATIM formula strings from the workbook's "General foods" row 3
     (energy/sat/sugar/salt/fibre/protein/fvl points, Points A/C, Score, letter)
     and the REAL "Scenario" threshold values.
  2. Stamp a broad grid of input products into a fresh workbook, replicating the
     original formulas row by row (openpyxl Translator shifts the relative cell
     refs, keeps the absolute Scenario!$X$n refs).
  3. Evaluate that workbook with the third-party `formulas` engine — Excel's own
     formula semantics, an independent code path from our Python engine.
  4. Compare our engine to the oracle on every grid point, especially the
     protein-cap region (negative points >= 11).

If our engine matches the oracle across hundreds of cases, the transcription of
BOTH the thresholds and the protein-cap rule is proven faithful.
"""
import sys, os, random, itertools

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import openpyxl
from openpyxl.formula.translate import Translator
import formulas
from nutriscore import score_general_food

ROOT = os.path.join(os.path.dirname(__file__), "..")
SRC_WB = os.path.join(ROOT, "data", "raw", "nutriscore_workbook.xlsx")
ORACLE_WB = os.path.join(ROOT, "data", "raw", "_oracle_grid.xlsx")

# Input columns on the General foods sheet (header row 2)
IN_COLS = dict(energy="B", sugar="C", sat="D", salt="F", fvl="H",
               fibre="I", protein="J")
# Formula columns we replicate verbatim from row 3
FORMULA_COLS = ["K", "M", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"]

LETTER = {"Nutriscore_A": "A", "Nutriscore_B": "B", "Nutriscore_C": "C",
          "Nutriscore_D": "D", "Nutriscore_E": "E"}

# Boundary-focused value sets (hit thresholds + just above/below)
GRID = dict(
    energy=[0, 335, 336, 670, 1675, 3350, 3351, 4000],
    sugar=[0, 3.4, 3.5, 6.8, 17, 34, 51, 52, 70],
    sat=[0, 1, 1.5, 5, 10, 11, 15],
    salt=[0, 0.2, 0.5, 2.0, 4.0, 4.1, 6],
    fibre=[0, 3, 4.1, 7.4, 8, 12],
    protein=[0, 2.4, 12, 17, 18, 30, 80],
    fvl=[0, 40, 60, 80, 100],
)


def build_cases(n_random=500, seed=42):
    """Per-nutrient boundary sweeps + a seeded random cross-sample."""
    rnd = random.Random(seed)
    base = dict(energy=1600, sugar=10, sat=3, salt=0.5, fibre=2, protein=10, fvl=0)
    cases = []
    # 1) sweep each nutrient through its full set, others at base
    for nutr, values in GRID.items():
        for v in values:
            c = dict(base); c[nutr] = v
            cases.append(c)
    # 2) random cross-product sample (naturally covers protein-cap region)
    for _ in range(n_random):
        cases.append({k: rnd.choice(v) for k, v in GRID.items()})
    # 3) explicit protein-cap stressors: high negatives + high protein
    for e, s, sf, sa, p in itertools.product(
            [3350, 4000], [34, 51, 70], [10, 15], [2.0, 4.0], [17, 30, 80]):
        cases.append(dict(energy=e, sugar=s, sat=sf, salt=sa,
                          fibre=0, protein=p, fvl=0))
    # de-dup
    seen, uniq = set(), []
    for c in cases:
        key = tuple(c[k] for k in ("energy", "sugar", "sat", "salt", "fibre", "protein", "fvl"))
        if key not in seen:
            seen.add(key); uniq.append(c)
    return uniq


def build_oracle_workbook(cases):
    """Fresh workbook: real Scenario values + verbatim formulas over the grid."""
    src = openpyxl.load_workbook(SRC_WB, data_only=False)
    gf = src["General foods"]
    row3 = {c: gf[f"{c}3"].value for c in FORMULA_COLS}

    out = openpyxl.Workbook()
    # copy Scenario sheet values verbatim
    sc_out = out.active
    sc_out.title = "Scenario"
    sc_src = openpyxl.load_workbook(SRC_WB, data_only=True)["Scenario"]
    for row in sc_src.iter_rows():
        for cell in row:
            if cell.value is not None:
                sc_out[cell.coordinate] = cell.value

    ws = out.create_sheet("General foods")
    for r, case in enumerate(cases, start=3):
        for nutr, col in IN_COLS.items():
            ws[f"{col}{r}"] = case[nutr]
        for col in FORMULA_COLS:
            ws[f"{col}{r}"] = Translator(row3[col], origin=f"{col}3").translate_formula(f"{col}{r}")
    out.save(ORACLE_WB)
    return len(cases)


def run():
    cases = build_cases()
    n = build_oracle_workbook(cases)
    print(f"grid cases: {n}  -> evaluating with `formulas` engine (independent oracle)...")

    xl = formulas.ExcelModel().loads(ORACLE_WB).finish()
    sol = xl.calculate()

    # build a lookup from solution keys -> value, for AD/AE cells
    def cell(coord):
        # keys look like "'[_ORACLE_GRID.XLSX]GENERAL FOODS'!AD3"
        for k, v in sol.items():
            if k.upper().endswith(f"!{coord}") and "GENERAL FOODS" in k.upper():
                val = v.value
                try:
                    return val[0, 0]
                except Exception:
                    return val
        return None

    checked = mismatch = 0
    fails = []
    for r, case in enumerate(cases, start=3):
        exp_score = cell(f"AD{r}")
        exp_letter_raw = cell(f"AE{r}")
        exp_letter = LETTER.get(exp_letter_raw, exp_letter_raw)
        res = score_general_food(
            energy_kj=case["energy"], sugar_g=case["sugar"], sat_fat_g=case["sat"],
            salt_g=case["salt"], fibre_g=case["fibre"], protein_g=case["protein"],
            fvl_percent=case["fvl"])
        checked += 1
        try:
            exp_score_i = int(round(float(exp_score)))
        except Exception:
            exp_score_i = exp_score
        if res.score != exp_score_i or res.letter != exp_letter:
            mismatch += 1
            if len(fails) < 25:
                fails.append(f"  case {case} -> engine {res.score}/{res.letter} "
                             f"| oracle {exp_score_i}/{exp_letter}")

    print(f"checked={checked}  mismatch={mismatch}")
    # how many exercised the protein-cap path
    capped = sum(1 for c in cases
                 if not score_general_food(c["energy"], c["sugar"], c["sat"], c["salt"],
                                           c["fibre"], c["protein"], c["fvl"]).protein_counted)
    print(f"protein-cap cases exercised: {capped}")
    if fails:
        print("MISMATCHES (first 25):")
        print("\n".join(fails))
    ok = mismatch == 0 and checked > 0
    print("\nRESULT:", "PASS  ENGINE == INDEPENDENT ORACLE" if ok else "FAIL")
    return ok


if __name__ == "__main__":
    sys.exit(0 if run() else 1)
