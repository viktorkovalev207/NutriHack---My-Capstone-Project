"""
GOLDEN TEST — prove the engine reproduces the official workbook exactly.

For every example product on the "General foods" sheet we feed the engine the
same inputs the workbook used and compare our result to the workbook's own
"Updated algorithm" output:
    expected raw score  -> column AD
    expected letter     -> column AE  (Nutriscore_X)

NOTE ON COVERAGE: this workbook is a TEMPLATE and ships only 3 filled example
products (apple sauce, sea salt, granulated sugar). So this test confirms the
engine on those 3 real reference points only — necessary but NOT sufficient.
The real proof of correctness is tests/oracle_validation.py, which checks the
engine against an independent formula-engine oracle over 620 grid cases
(534 in the protein-cap region). Keep both: this pins the 3 official examples,
the oracle pins the full value range.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import openpyxl
from nutriscore import score_general_food

WB = os.path.join(os.path.dirname(__file__), "..", "data", "raw", "nutriscore_workbook.xlsx")

# General foods column map (per sheet header row 2)
COL = dict(name="A", energy="B", sugar="C", sat="D", salt="F", fvl="H",
           fibre="I", protein="J", exp_score="AD", exp_letter="AE")

LETTER = {"Nutriscore_A": "A", "Nutriscore_B": "B", "Nutriscore_C": "C",
          "Nutriscore_D": "D", "Nutriscore_E": "E"}


def run():
    ws = openpyxl.load_workbook(WB, data_only=True)["General foods"]
    checked = skipped = failed = 0
    failures = []
    for r in range(3, ws.max_row + 1):
        def g(key):
            return ws[f"{COL[key]}{r}"].value
        name = g("name")
        if name is None:
            continue
        vals = [g(k) for k in ("energy", "sugar", "sat", "salt", "fvl", "fibre", "protein")]
        exp_score, exp_letter_raw = g("exp_score"), g("exp_letter")
        # workbook flags incomplete rows as "missing data" -> nothing to check
        if exp_letter_raw == "missing data" or exp_letter_raw is None:
            skipped += 1
            continue
        if any(v is None for v in vals):
            skipped += 1
            continue
        energy, sugar, sat, salt, fvl, fibre, protein = vals
        res = score_general_food(energy_kj=energy, sugar_g=sugar, sat_fat_g=sat,
                                 salt_g=salt, fibre_g=fibre, protein_g=protein,
                                 fvl_percent=fvl)
        exp_letter = LETTER.get(exp_letter_raw, exp_letter_raw)
        checked += 1
        if res.score != exp_score or res.letter != exp_letter:
            failed += 1
            if len(failures) < 20:
                failures.append(
                    f"  row {r:>4} {str(name)[:28]:28} "
                    f"got score={res.score:>3} {res.letter} | "
                    f"exp score={exp_score:>3} {exp_letter}")

    print(f"checked={checked}  skipped={skipped}  failed={failed}")
    if failures:
        print("FAILURES (first 20):")
        print("\n".join(failures))
    ok = failed == 0 and checked > 0
    print("\nRESULT:", "PASS  ENGINE MATCHES OFFICIAL WORKBOOK" if ok else "FAIL")
    return ok


if __name__ == "__main__":
    sys.exit(0 if run() else 1)
